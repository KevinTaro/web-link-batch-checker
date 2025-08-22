def process_xlsx_check(xlsx_file, checked_dir='output/checked'):
    """
    檢查 Excel 檔案中的網址，並將結果儲存到 checked_dir
    """
    wb = load_workbook(xlsx_file)
    for ws in wb.worksheets:
        print(f'正在檢查工作表：{ws.title}')
        header_row = 2 if ws.cell(row=2, column=1).value == '標題' else 1
        url_col = None
        for col in range(1, ws.max_column+1):
            if ws.cell(row=header_row, column=col).value == '網址':
                url_col = col
                break
        if not url_col:
            print(f'找不到「網址」欄位，略過工作表 {ws.title}')
            continue
        result_col = ws.max_column + 1
        ws.cell(row=header_row, column=result_col, value='網址檢查結果')
        urls = []
        row_url_map = {}
        for row in range(header_row+1, ws.max_row+1):
            url = ws.cell(row=row, column=url_col).value
            if url:
                urls.append(url)
                row_url_map[row] = url
        results = asyncio.run(batch_check_urls(urls, timeout=5, retries=3))
        for row, url in row_url_map.items():
            ws.cell(row=row, column=result_col, value=results.get(url, ''))
    out_path = os.path.join(checked_dir, f'檢查結果_{os.path.basename(xlsx_file)}')
    wb.save(out_path)
    print(f'已完成檢查，結果儲存為 {out_path}')

def cli_all():
    while True:
        print('=== 一鍵執行完整流程 (CLI) ===')
        print('請輸入要批次抓取的網址，每行一個，輸入 quit 結束輸入：')
        urls = []
        while True:
            url = input().strip()
            if url.lower() == 'quit':
                break
            if url:
                if not url.startswith(('http://', 'https://')):
                    url = 'https://' + url
                urls.append(url)
        if not urls:
            print('未輸入任何網址，流程結束。')
            break
        import web_grab_tool
        print('正在批次抓取...')
        web_grab_tool.gui_main(urls)
        print('網頁連結已批次抓取並匯出至 output/csv 及 output/xlsx')
        # 自動選擇最新 xlsx 檔案
        xlsx_dir = os.path.join('output', 'xlsx')
        xlsx_files = [f for f in os.listdir(xlsx_dir) if f.endswith('.xlsx') and f.startswith('打包網頁連結_')]
        if not xlsx_files:
            print('找不到要檢查的 xlsx 檔案')
            continue
        xlsx_files.sort(reverse=True)
        xlsx_file = os.path.join(xlsx_dir, xlsx_files[0])
        print('正在檢查網址...')
        process_xlsx_check(xlsx_file)
        print('檢查結果已儲存於 output/checked')
        print('--- 流程結束，可繼續輸入網址或輸入 quit 離開 ---')
import pandas as pd
import aiohttp
import asyncio
from openpyxl import load_workbook
import os



async def check_url(session, url, timeout=5, retries=2):
    for attempt in range(retries):
        try:
            async with session.get(url, timeout=timeout) as resp:
                if resp.status == 200:
                    return 'OK'
                else:
                    return f'HTTP {resp.status}'
        except asyncio.TimeoutError:
            if attempt < retries - 1:
                timeout *= 2  # 慢網站自動延長 timeout
                continue
            return 'Timeout'
        except aiohttp.ClientError as e:
            if attempt < retries - 1:
                await asyncio.sleep(1)
                continue
            return str(e)
        except Exception as e:
            return str(e)


async def batch_check_urls(urls, timeout=5, retries=2):
    results = {}
    connector = aiohttp.TCPConnector(limit=64, ssl=False)
    async with aiohttp.ClientSession(connector=connector) as session:
        tasks = [check_url(session, url, timeout=timeout, retries=retries) for url in urls]
        responses = await asyncio.gather(*tasks)
        for url, result in zip(urls, responses):
            results[url] = result
    return results



def cli_main():
    xlsx_file = input('請輸入要檢查的 xlsx 檔案名稱：').strip()
    if not os.path.isabs(xlsx_file):
        xlsx_file = os.path.join('output', 'xlsx', xlsx_file)
    process_xlsx_check(xlsx_file)

def gui_main(xlsx_file):
    process_xlsx_check(xlsx_file)


def run_project():
    print("=== 網頁批次抓取與網址檢查專案 ===")
    import web_grab_tool
    import asyncio
    asyncio.run(web_grab_tool.main())
    xlsx_files = [f for f in os.listdir(os.path.join('output', 'xlsx')) if f.endswith('.xlsx') and f.startswith('打包網頁連結_')]
    if not xlsx_files:
        print('找不到要檢查的 xlsx 檔案')
        return
    xlsx_files.sort(reverse=True)
    xlsx_file = os.path.join('output', 'xlsx', xlsx_files[0])
    print(f'自動選擇最新檔案：{xlsx_file}')
    process_xlsx_check(xlsx_file)

if __name__ == '__main__':
    print('請選擇執行模式：')
    print('1. 只檢查 Excel 網址 (原CLI)')
    print('2. 一鍵執行完整流程 (抓取+檢查)')
    mode = input('請輸入 1 或 2：').strip()
    if mode == '2':
        cli_all()
    else:
        cli_main()
