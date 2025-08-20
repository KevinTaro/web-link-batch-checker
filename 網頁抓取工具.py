import requests
from bs4 import BeautifulSoup
import csv
import urllib.parse
from datetime import datetime
import os
import glob
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed

def get_webpage_links(url):
    """
    抓取網頁上的所有標題和連結
    """
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        h2_titles = [h2.get_text(strip=True) for h2 in soup.find_all('h2', class_='Index_title')]
        links = soup.find_all('a')
        data = []
        for link in links:
            a_text = link.get_text(strip=True)
            href = link.get('href')
            a_title = link.get('title') if link.has_attr('title') else ''
            target_blank = '是' if link.get('target') == '_blank' else ''
            if href and not href.strip().lower().startswith('javascript:'):
                absolute_url = urllib.parse.urljoin(url, href)
                if not a_text:
                    a_text = href
                index_title = h2_titles[0] if h2_titles else ''
                data.append([a_text, absolute_url, a_title, target_blank, index_title])
        for h2_title in h2_titles:
            data.append([h2_title, '', '', '', h2_title])
        return data
    except requests.exceptions.RequestException as e:
        print(f"網頁請求錯誤: {e}")
        return []
    except Exception as e:
        print(f"處理過程發生錯誤: {e}")
        return []

def save_to_csv(data, filename):
    """
    將資料儲存為 CSV 檔案
    """
    try:
        # 取得網頁名稱、原始網址、title 文字
        page_name = ''
        for row in data:
            if row[4]:
                page_name = row[4]
                break
        if not page_name:
            page_name = filename.split('_')[1] if '_' in filename else filename
        # 網址取 filename 裡的 domain
        page_url = ''
        if '_' in filename:
            page_url = filename.split('_')[2].replace('.csv','')
        # 取得原始輸入網址
        input_url = page_url if page_url.startswith('http') else ''
        # 取得 <title> 內容
        title_text = ''
        try:
            import requests
            from bs4 import BeautifulSoup
            resp = requests.get(input_url, timeout=10)
            soup = BeautifulSoup(resp.text, 'html.parser')
            title_tag = soup.find('title')
            if title_tag:
                title_text = title_tag.get_text(strip=True)
        except Exception:
            pass
        with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            # 第一列：網頁名稱、title 文字、網頁網址、輸入網址
            writer.writerow(["網頁名稱", title_text, "網頁網址", input_url])
            # 欄位加上「日期戳記」
            writer.writerow(['標題', '網址', 'title屬性', '另開新視窗', 'Index_title', '日期戳記'])
            for row in data:
                import re
                date_str = ''
                for col in row[:2]:
                    m = re.search(r'(20\d{2}[/-]?\d{2}[/-]?\d{2})', col)
                    if m:
                        date_str = m.group(1)
                        break
                writer.writerow(row + [date_str])
        print(f"已成功儲存 {len(data)} 筆資料到 {filename}")
        return True
    except Exception as e:
        print(f"儲存檔案時發生錯誤: {e}")
        return False

def main():
    """
    主程式
    """
    print("=== 網頁連結抓取工具 ===")
    print("此工具可以抓取網頁上的所有連結和標題，並儲存為 CSV 檔案")
    print()

    csv_files = []
    print("請輸入要抓取的多個網頁網址，每行一個，輸入 'quit' 結束輸入：")
    urls = []
    while True:
        url = input().strip()
        if url.lower() == 'quit':
            break
        if url:
            urls.append(url)

    # 多執行緒批次抓取所有網址
    def fetch_one(url):
        if not url.startswith(('http://', 'https://')):
            url = 'https://' + url
        data = get_webpage_links(url)
        return url, data

    results = {}
    with ThreadPoolExecutor(max_workers=8) as executor:
        future_to_url = {executor.submit(fetch_one, url): url for url in urls}
        for future in as_completed(future_to_url):
            url, data = future.result()
            if not data:
                print(f"{url} 沒有找到任何連結，或抓取失敗")
                continue
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            domain = urllib.parse.urlparse(url).netloc
            filename = os.path.join('output', 'csv', f"網頁連結_{domain}_{timestamp}.csv")
            if save_to_csv(data, filename):
                print(f"檔案已儲存為: {filename}")
                print("可以用 Excel 開啟此檔案")
                csv_files.append(filename)
            print("-" * 50)

    # 結束時打包成 xlsx
    if csv_files:
        xlsx_name = os.path.join('output', 'xlsx', f"打包網頁連結_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        from openpyxl.utils import get_column_letter
        from openpyxl import load_workbook
        with pd.ExcelWriter(xlsx_name, engine='openpyxl') as writer:
            for csv_file in csv_files:
                try:
                    with open(csv_file, encoding='utf-8-sig') as f:
                        lines = f.readlines()
                        info_row = [cell.strip() for cell in lines[0].split(',')]
                    df = pd.read_csv(csv_file, skiprows=1)
                    sheet_name = str(info_row[1])[:31] if info_row[1] else os.path.basename(csv_file).split('_')[1]
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                except Exception as e:
                    print(f"匯入 {csv_file} 到 Excel 時發生錯誤: {e}")
        from openpyxl.utils import get_column_letter
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_name)
        for ws in wb.worksheets:
            info = None
            for csv_file in csv_files:
                with open(csv_file, encoding='utf-8-sig') as f:
                    info_row = [cell.strip() for cell in f.readline().split(',')]
                sheet_name = str(info_row[1])[:31] if info_row[1] else os.path.basename(csv_file).split('_')[1]
                if ws.title == sheet_name:
                    info = info_row
                    break
            if info:
                ws.insert_rows(1)
                for i, val in enumerate(info):
                    ws.cell(row=1, column=i+1, value=val)
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        wb.save(xlsx_name)
    print(f"所有網頁已打包成 Excel 檔案：{xlsx_name}")

if __name__ == "__main__":
    main()
