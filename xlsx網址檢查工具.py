import pandas as pd
import aiohttp
import asyncio
from openpyxl import load_workbook
import os



async def check_url(session, url, timeout=5, retries=2):
    for attempt in range(retries):
        try:
            async with session.get(url, timeout=timeout) as resp:
                if resp.status == 200:
                    return 'OK'
                else:
                    return f'HTTP {resp.status}'
        except asyncio.TimeoutError:
            if attempt < retries - 1:
                timeout *= 2  # 慢網站自動延長 timeout
                continue
            return 'Timeout'
        except aiohttp.ClientError as e:
            if attempt < retries - 1:
                await asyncio.sleep(1)
                continue
            return str(e)
        except Exception as e:
            return str(e)


async def batch_check_urls(urls, timeout=5, retries=2):
    results = {}
    connector = aiohttp.TCPConnector(limit=64, ssl=False)
    async with aiohttp.ClientSession(connector=connector) as session:
        tasks = [check_url(session, url, timeout=timeout, retries=retries) for url in urls]
        responses = await asyncio.gather(*tasks)
        for url, result in zip(urls, responses):
            results[url] = result
    return results


def main():
    xlsx_file = input('請輸入要檢查的 xlsx 檔案名稱：').strip()
    if not os.path.isabs(xlsx_file):
        xlsx_file = os.path.join('output', 'xlsx', xlsx_file)
    wb = load_workbook(xlsx_file)
    for ws in wb.worksheets:
        print(f'正在檢查工作表：{ws.title}')
        header_row = 2 if ws.cell(row=2, column=1).value == '標題' else 1
        url_col = None
        for col in range(1, ws.max_column+1):
            if ws.cell(row=header_row, column=col).value == '網址':
                url_col = col
                break
        if not url_col:
            print(f'找不到「網址」欄位，略過工作表 {ws.title}')
            continue
        result_col = ws.max_column + 1
        ws.cell(row=header_row, column=result_col, value='網址檢查結果')
        urls = []
        row_url_map = {}
        for row in range(header_row+1, ws.max_row+1):
            url = ws.cell(row=row, column=url_col).value
            if url:
                urls.append(url)
                row_url_map[row] = url
        # async 批次檢查
        results = asyncio.run(batch_check_urls(urls, timeout=5, retries=3))
        for row, url in row_url_map.items():
            ws.cell(row=row, column=result_col, value=results.get(url, ''))
    out_path = os.path.join('output', 'checked', f'檢查結果_{os.path.basename(xlsx_file)}')
    wb.save(out_path)
    print(f'已完成檢查，結果儲存為 {out_path}')


def run_project():
    print("=== 網頁批次抓取與網址檢查專案 ===")
    import 網頁抓取工具
    import asyncio
    asyncio.run(網頁抓取工具.main())
    xlsx_files = [f for f in os.listdir(os.path.join('output', 'xlsx')) if f.endswith('.xlsx') and f.startswith('打包網頁連結_')]
    if not xlsx_files:
        print('找不到要檢查的 xlsx 檔案')
        return
    xlsx_files.sort(reverse=True)
    xlsx_file = xlsx_files[0]
    print(f'自動選擇最新檔案：{xlsx_file}')
    wb = load_workbook(os.path.join('output', 'xlsx', xlsx_file))
    for ws in wb.worksheets:
        print(f'正在檢查工作表：{ws.title}')
        header_row = 2 if ws.cell(row=2, column=1).value == '標題' else 1
        url_col = None
        for col in range(1, ws.max_column+1):
            if ws.cell(row=header_row, column=col).value == '網址':
                url_col = col
                break
        if not url_col:
            print(f'找不到「網址」欄位，略過工作表 {ws.title}')
            continue
        result_col = ws.max_column + 1
        ws.cell(row=header_row, column=result_col, value='網址檢查結果')
        urls = []
        row_url_map = {}
        for row in range(header_row+1, ws.max_row+1):
            url = ws.cell(row=row, column=url_col).value
            if url:
                urls.append(url)
                row_url_map[row] = url
        results = asyncio.run(batch_check_urls(urls, timeout=5, retries=3))
        for row, url in row_url_map.items():
            ws.cell(row=row, column=result_col, value=results.get(url, ''))
    out_path = os.path.join('output', 'checked', f'檢查結果_{xlsx_file}')
    wb.save(out_path)
    print(f'已完成檢查，結果儲存為 {out_path}')

if __name__ == '__main__':
    run_project()
